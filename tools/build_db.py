# REDHOOD 게임 데이터 종합본 생성기
# 실행: python3 tools/build_db.py  (data/*.json 을 읽어 docs/REDHOOD_game_db.xlsx 을 새로 씀)
# 판을 올릴 때마다 다시 돌리면 표가 코드와 어긋나지 않는다.
import json, datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

D = lambda n: json.load(open(f'/home/claude/redhood/data/{n}.json'))
scoring, dice, relics, enemies = D('scoring'), D('dice'), D('relics'), D('enemies')
events, acts, act1 = D('events'), D('acts'), D('act1')
statuses = D('statuses')

EN = {e['id']: e for e in enemies}
VERSION = 'v2.16'
TODAY = '2026-08-08'

FONT = 'Arial'
HEAD_FILL = PatternFill('solid', fgColor='3A2A1C')
SUB_FILL  = PatternFill('solid', fgColor='EDE3D2')
TITLE_F   = Font(name=FONT, size=15, bold=True, color='3A2A1C')
HEAD_F    = Font(name=FONT, size=10, bold=True, color='F4EAD5')
BODY_F    = Font(name=FONT, size=10)
NOTE_F    = Font(name=FONT, size=9, italic=True, color='7A6850')
INPUT_F   = Font(name=FONT, size=10, color='0000FF')
THIN = Side(style='thin', color='CFC3AE')
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TIER_FILL = {'common': None,
             'uncommon': PatternFill('solid', fgColor='E4EDF5'),
             'rare': PatternFill('solid', fgColor='ECE4F5'),
             'epic': PatternFill('solid', fgColor='FBF0DA'),
             'legendary': PatternFill('solid', fgColor='FBF0DA'),
             'normal': None, 'elite': PatternFill('solid', fgColor='E4EDF5'),
             'boss': PatternFill('solid', fgColor='F8E2E0')}

wb = openpyxl.Workbook()
wb.remove(wb.active)

def sheet(name, title, note, headers, rows, widths, freeze='A4'):
    ws = wb.create_sheet(name)
    ws['A1'] = title; ws['A1'].font = TITLE_F
    ws['A2'] = note;  ws['A2'].font = NOTE_F
    for i, hdr in enumerate(headers, 1):
        c = ws.cell(row=3, column=i, value=hdr)
        c.font = HEAD_F; c.fill = HEAD_FILL; c.border = BOX
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for r, row in enumerate(rows, 4):
        tier = row[-1] if isinstance(row[-1], str) and row[-1] in TIER_FILL else None
        for i, v in enumerate(row[:len(headers)], 1):
            c = ws.cell(row=r, column=i, value=v)
            c.font = BODY_F; c.border = BOX
            c.alignment = Alignment(vertical='top', wrap_text=isinstance(v, str) and len(str(v)) > 26)
            if tier and TIER_FILL.get(tier): c.fill = TIER_FILL[tier]
    for i, wd in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = wd
    ws.row_dimensions[3].height = 26
    ws.freeze_panes = freeze
    return ws

# ---------- 효과 사람말 ----------
OPNAME = {'rest':'휴식','selfDamage':'자해','poison':'독','bleed':'출혈','damage':'피해','block':'방어','empower':'힘','confuse':'혼란','heal':'회복',
          'weaken':'약화','vulnerable':'취약','bleed':'출혈','regen':'재생','focus':'집중',
          'strength':'힘','loseHp':'HP 감소','gainRelic':'유물 획득','drain':'흡혈',
          'markReroll':'리롤 표식','curse':'저주'}
ST_NAME = {}; ST_AMT = {}; ST_TURN = {}
NO_POWER = {'onUseFaceDamage','onUseFaceCoin','noReroll','zeroValue','hideFace','needReroll','linked','spread','faceLow','faceHigh','rerollCost'}
ST_RULE = {}
def eff_text(effs):
    if not effs: return '—'
    out = []
    for e in effs:
        if e.get('op') == 'status':
            pw = 0 if ST_RULE.get(e.get('kind')) in NO_POWER else (e.get('power') or ST_AMT.get(e.get('kind'), 0))
            t = f"{ST_NAME.get(e.get('kind'), e.get('kind'))}"
            if pw: t += f" 폭발{pw}"
            t += f" {e.get('amount', 1)}칸"
            out.append(t)
            continue
        n = OPNAME.get(e.get('op'), e.get('op'))
        a = e.get('amount')
        h = e.get('hits')
        txt = n if (a is None or e.get('op') == 'rest') else f"{n} {a}"
        if h and h > 1: txt += f" ×{h}타"
        out.append(txt)
    return ' · '.join(out)

ST_NAME.update({x['id']: x['name'] for x in statuses['list']})
ST_AMT.update({x['id']: x.get('amount', 0) for x in statuses['list']})
ST_TURN.update({x['id']: x.get('turns', 0) for x in statuses['list']})
ST_RULE.update({x['id']: x['rule'] for x in statuses['list']})

ACT_OF = {}
for a in acts['acts']:
    for t in a['themes']:
        for k in ('normals','elites'):
            for e in t.get(k, []): ACT_OF.setdefault(e, (a['act'], t['name']))
        ACT_OF.setdefault(t['boss'], (a['act'], t['name']))
ACT_OF.setdefault(acts['finalBoss'], (4, '최후의 어둠'))

# 표 순서: 막 → 테마 → 난이도(일반·정예·보스). 지도에 나오는 차례 그대로 훑을 수 있게.
ORD = {}
for a in acts['acts']:
    for t in a['themes']:
        for k in ('normals', 'elites'):
            for e in t.get(k, []): ORD.setdefault(e, len(ORD))
        ORD.setdefault(t['boss'], len(ORD))
ORD.setdefault(acts['finalBoss'], len(ORD))
ENEMIES = sorted(enemies, key=lambda e: ORD.get(e['id'], 9999))

# ---------- 1. 개요 ----------
ws = wb.create_sheet('개요')
ws['A1'] = 'REDHOOD 게임 데이터 종합본'; ws['A1'].font = Font(name=FONT, size=18, bold=True, color='8F2A20')
rows = [
 ('판', VERSION), ('갱신일', TODAY),
 ('출처', 'data/*.json 에서 자동 생성 — 손으로 고치지 말고 JSON을 고친 뒤 다시 뽑을 것'),
 ('', ''),
 ('시작 HP', act1['player']['maxHp']),
 ('시작 주사위', ', '.join(act1['player']['startDice'])),
 ('턴당 리롤', scoring['rerollsPerTurn']),
 ('한 막의 층수', act1['map']['floors']),
 ('층당 적 HP 증가', act1['hpScalePerFloor']),
 ('막별 HP 배율', ' / '.join(f"{k}막 ×{v}" for k, v in acts['scaling']['hp'].items())),
 ('막별 공격 배율', ' / '.join(f"{k}막 ×{v}" for k, v in acts['scaling']['atk'].items())),
 ('행동 발동 조건', '가중치(뽑힐 확률·0이면 연계 전용) · 해금 턴 · 락 턴 · 쿨다운 · 연계 — 다섯 가지뿐'),
 ('강화 행동', "별도 장치가 아니다. 각 적의 'surge' 는 해금 턴·쿨다운이 붙은 평범한 행동일 뿐"),
 ('표 정렬', '적·적행동 시트는 막 → 테마 → 난이도(일반·정예·보스) 순'),
 ('보스 처치 회복', f"최대 HP의 {int(acts['bossHealRatio']*100)}%"),
 ('최종 보스', EN[acts['finalBoss']]['name']),
 ('', ''),
 ('족보', f"{len(scoring['categories'])}종 · 변형 {sum(len(c['variants']) for c in scoring['categories'])}개"),
 ('주사위', f"{len(dice)}종"), ('유물', f"{len(relics)}개"),
 ('적', f"{len(enemies)}종 (일반 {sum(1 for e in enemies if e['tier']=='normal')} · 정예 {sum(1 for e in enemies if e['tier']=='elite')} · 보스 {sum(1 for e in enemies if e['tier']=='boss')})"),
 ('무기', f"{len(events['weapons'])}종"), ('만남 이벤트', f"{len(events['events'])}개"),
]
for r, (k, v) in enumerate(rows, 3):
    a = ws.cell(row=r, column=1, value=k); a.font = Font(name=FONT, size=10, bold=True)
    b = ws.cell(row=r, column=2, value=v); b.font = BODY_F
    b.alignment = Alignment(wrap_text=True, vertical='top')
ws.column_dimensions['A'].width = 20; ws.column_dimensions['B'].width = 74
r0 = len(rows) + 5
ws.cell(row=r0, column=1, value='시트 안내').font = TITLE_F
for i, (n, d) in enumerate([
  ('막·테마','3막 × 테마 3종. 어느 테마에 어느 적이 나오는지'),
  ('족보','9종 × 변형 18개. 성립 조건·피해식·부가 능력'),
  ('주사위','13종의 면 구성과 등급'),
  ('유물','26개. 일반 17 / 정예 9'),
  ('적','44종. HP·등장 위치·강화 행동·연계·해금 턴'),
  ('적행동','모든 행동 낱개. 효과 수치와 발동 조건'),
  ('무기','6종과 시작 족보'),
  ('만남','11개 이벤트의 선택지와 대가'),
  ('지도·보상','노드 생성 규칙과 보상 확률'),
  ('경제','코인·상점 가격·족보 교체 비용'),
  ('계몽','20단계 난이도 상승'),
], 1):
    ws.cell(row=r0+i, column=1, value=n).font = Font(name=FONT, size=10, bold=True)
    ws.cell(row=r0+i, column=2, value=d).font = BODY_F
ws.freeze_panes = 'A3'

# ---------- 2. 막·테마 ----------
rows = []
for a in acts['acts']:
    for t in a['themes']:
        rows.append([a['act'], t['name'], t['id'],
                     ', '.join(EN[x]['name'] for x in t.get('normals', [])),
                     ', '.join(EN[x]['name'] for x in t.get('elites', [])),
                     EN[t['boss']]['name'],
                     ', '.join(t.get('events', [])),
                     f"bg_{t['id']}.jpg"])
rows.append([4, '최후의 어둠', 'final', '—', '—', EN[acts['finalBoss']]['name'], '—', 'bg_final.jpg'])
sheet('막·테마', '막 · 테마', '런마다 막당 테마 1개가 무작위로 뽑힌다. 보스는 테마 고정, 일반·정예·이벤트는 그 테마 풀 안에서 랜덤.',
      ['막','테마','id','일반 적','정예','보스','이벤트','배경 파일'], rows, [6,14,10,42,18,16,24,16])

# ---------- 3. 족보 ----------
rows = []
for c in scoring['categories']:
    for v in c['variants']:
        rows.append([c['id'], c['name'], v['id'], v['name'], v.get('tier','common'),
                     c.get('ruleText',''), str(c.get('score','')),
                     v.get('abilityText','') or '부가 없음',
                     ' '.join(str(x) for x in c.get('example',[])),
                     c.get('fx',''), v.get('tier','common')])
sheet('족보', '족보 9종 · 변형 18개', '같은 족보라도 변형에 따라 부가 능력과 등급이 다르다. 등급은 색으로 구분(흰 커먼 / 파랑 언커먼 / 보라 레어 / 금 전설).',
      ['족보 id','족보','변형 id','변형 이름','등급','성립 조건','피해식','부가 능력','예시','연출'],
      rows, [12,12,16,18,9,40,14,44,12,9])

# ---------- 4. 주사위 ----------
rows = [[d['id'], d['name'], ' '.join(str(f) for f in d['faces']),
         d.get('tier','common'), d.get('desc',''), d.get('tier','common')] for d in dice]
sheet('주사위', '주사위 13종', '면 구성이 곧 성능이다. 상점과 전투 보상에서 얻어 5개 중 하나를 교체한다.',
      ['id','이름','면 구성','등급','설명'], rows, [12,16,18,9,50])

# ---------- 5. 유물 ----------
rows = [[r['id'], r['name'], r.get('icon',''), '정예' if r.get('tier')=='elite' else '일반',
         r['desc'], r.get('tier','normal')] for r in relics]
sheet('유물', '유물 26개', '일반 17개는 어디서나, 정예 9개는 정예 처치 보상에서만. 정예는 항상 유물을 떨구고 그중 10% 확률로 정예 유물이 나온다.',
      ['id','이름','아이콘','등급','효과'], rows, [18,16,7,8,58])

# ---------- 6. 적 ----------
rows = []
for e in ENEMIES:
    act, theme = ACT_OF.get(e['id'], ('—','—'))
    hp = e['hp']; hp_s = f"{hp[0]}" if hp[0]==hp[1] else f"{hp[0]}~{hp[1]}"
    late = [m['name'] for m in e['moves'].values() if m.get('minTurn')]
    lateT = [m['minTurn'] for m in e['moves'].values() if m.get('minTurn')]
    chains = []
    for mid, m in e['moves'].items():
        if m.get('followUp'):
            fu = m['followUp']
            chains.append(f"{m['name']} → {e['moves'][fu['move']]['name']} ({int(fu['chance']*100)}%)")
    phases = len(e.get('phases', [])) or 1
    rows.append([e['id'], e['name'], {'normal':'일반','elite':'정예','boss':'보스'}[e['tier']],
                 act, theme, hp_s, len(e['moves']), phases,
                 (e.get('defaultMove') or '—'),
                 ', '.join((e.get('uniqueMoves') or {}).keys()) or '—',
                 (f"{late[0]} ({lateT[0]}턴부터)" if late else '—'),
                 ' / '.join(chains) if chains else '—',
                 e.get('enlightenedMove',{}).get('name','—'),
                 e['tier']])
sheet('적', '적 44종', '막 → 테마 → 난이도 순으로 늘어놓았다. 행동 발동 조건은 가중치·해금 턴·락 턴·쿨다운·연계 다섯 가지다.',
      ['id','이름','격','막','테마','HP','행동 수','국면','기본 행동','유니크 행동','해금 행동','연계','계몽 전용 행동'],
      rows, [16,16,6,5,14,10,7,6,16,20,22,34,18])

# ---------- 7. 적 행동 ----------
def weight_text(e, mid):
    pats = [(f"국면{i+1}", p['pattern']) for i, p in enumerate(e['phases'])] if e.get('phases') \
        else ([('', e['pattern'])] if e.get('pattern') else [])
    out = []
    for label, p in pats:
        pre = (label + ' ') if label else ''
        if p['mode'] == 'sequence':
            i = p['order'].index(mid) + 1 if mid in p['order'] else 0
            out.append(pre + (f"순서 {i}번째" if i else '순서 밖'))
        else:
            w = p.get('weights', {})
            out.append(pre + (str(w[mid]) if mid in w else '없음'))
    return ' / '.join(out) if out else '—'

rows = []
for e in ENEMIES:
    for mid, m in e['moves'].items():
        fu = m.get('followUp')
        br = m.get('break')
        rows.append([e['id'], e['name'], mid, m['name'], eff_text(m.get('effects')),
                     weight_text(e, mid), m.get('minTurn') or '', m.get('lockTurn') or '', m.get('cooldown') or 0,
                     (f"{br['damage']} → {br['move']}" if br else ''), '숨김' if m.get('hidden') else '',
                     (f"{e['moves'][fu['move']]['name']} {int(fu['chance']*100)}%" if fu else ''),
                     e['tier']])
    for mid, m in (e.get('uniqueMoves') or {}).items():
        rows.append([e['id'], e['name'], mid, m['name'], eff_text(m.get('effects')),
                     '유니크(추첨 안 함)', '', '', '', '파쇄 안 됨', '', '', e['tier']])
    if e.get('enlightenedMove'):
        rows.append([e['id'], e['name'], '__enlight', e['enlightenedMove']['name'],
                     eff_text(e['enlightenedMove'].get('effects')), '3번째 행동마다', '', '', '', '', '계몽 17~19단계', '', e['tier']])
sheet('적행동', '적 행동 전량',
      '상태이상 부여는 [종류 N칸] 으로 읽는다 — 적마다 다를 수 있는 건 부패의 폭발 피해뿐이고, 지속 턴을 포함한 나머지는 상태이상 시트의 값으로 고정이다. 막 → 테마 → 난이도 순. 발동 조건 다섯 가지 — 가중치는 뽑힐 확률(0이면 추첨에서 빠지고 연계로만 나온다), 해금 턴은 "이 턴부터", 락 턴은 "이 턴이 되면 더 안 씀", 쿨다운은 "쓰고 나서 몇 턴 쉰다"(0=제한 없음, 2면 한 칸 걸러 나온다), 연계는 "앞 행동 다음에 확률로 확정".',
      ['적 id','적 이름','행동 id','행동명','효과','가중치','해금 턴','락 턴','쿨다운','파쇄','비고','연계'], rows, [16,16,16,20,30,18,8,8,8,18,14,22])

# ---------- 7-2. 상태이상 ----------
RULE_TEXT = {'onUseFaceDamage':'쓰면 눈금만큼 피해','onUseFaceCoin':'쓰면 눈금만큼 코인','noReroll':'다시 못 굴림',
 'zeroValue':'눈금 0으로 계산','faceLow':'수치 이하만 나옴','faceHigh':'수치 이상만 나옴','hideFace':'눈 가림',
 'needReroll':'한 번 굴리기 전 값 없음','fuse':'턴 지나면 폭발','linked':'같이 굴러감',
 'rerollCost':'리롤 더 소모','spread':'양옆으로 번짐'}
NO_VALUE = {'onUseFaceDamage','onUseFaceCoin','noReroll','zeroValue','hideFace','needReroll','linked','spread'}
rows = [[s['id'], s['name'], RULE_TEXT.get(s['rule'], s['rule']),
         ('—' if s['rule'] in NO_VALUE else s.get('amount', 0)),
         (s.get('turns') or '영구'), s.get('color',''), s.get('text','')] for s in statuses['list']]
sheet('상태이상', f"주사위 상태이상 {len(statuses['list'])}종",
      '주사위 한 칸에 하나씩 붙는다. 새로 걸면 빈 칸을 먼저 채우고, 다 차면 아무 칸이나 덮는다. '
      '지속 턴이 영구면 정화하거나 조건을 만족해야 풀린다. 적 행동에서 수치를 덮어쓸 수 있는 건 부패뿐이다.',
      ['id','이름','규칙','수치','지속 턴','색','설명'], rows, [12,10,24,7,8,10,52])

# ---------- 8. 무기 ----------
NAMEOF = {v['id']: (c['name'], v['name']) for c in scoring['categories'] for v in c['variants']}
rows = []
for w_ in events['weapons']:
    st = w_['start']
    rows.append([w_['id'], w_['name'], w_.get('icon',''), w_.get('desc',''),
                 ' · '.join(f"{NAMEOF[v][0]}={NAMEOF[v][1]}" for v in st.values() if v in NAMEOF)])
sheet('무기', '무기 6종', '무기를 고르면 시작 족보 3개가 정해진다. 나머지 족보는 런 중에 보상으로 채운다.',
      ['id','이름','아이콘','컨셉','시작 족보'], rows, [12,16,7,26,60])

# ---------- 9. 만남 ----------
rows = []
for ev in events['events']:
    for i, ch in enumerate(ev['choices']):
        rows.append([ev['id'] if i==0 else '', ev['npc']['name'] if i==0 else '',
                     ', '.join(ev.get('themes',[])) if i==0 else '',
                     ch['text'], ch.get('sub',''), eff_text(ch.get('effects')), ch.get('result','')])
sheet('만남', '만남 이벤트 11개', '테마에 맞는 이벤트만 등장한다. 계몽 15단계부터 대가가 1.5배로 가혹해진다.',
      ['이벤트 id','NPC','등장 테마','선택지','요약','효과','결과 대사'], rows, [16,16,20,26,22,22,44])

# ---------- 10. 지도·보상 ----------
mp = act1['map']; rw = act1['rewards']
rows = [
 ['층수', mp['floors'], '마지막 층이 보스'],
 ['한 층의 갈림길', f"{mp['choicesMin']}~{mp['choicesMax']}", '최대 4열까지만 사용'],
 ['고정 층', ' / '.join(f"{k}층={v}" for k,v in mp['fixed'].items()), '보스 직전 층은 항상 휴식'],
 ['정예 등장 층', f"{mp['eliteFloors'][0]}~{mp['eliteFloors'][-1]}층", ''],
 ['노드 비중', ' / '.join(f"{k} {v}" for k,v in mp['nodeWeights'].items()), '적3 : 이벤트1 : 정예1.6 : 상점0.8 : 휴식0.5'],
 ['규칙 1', '연속 휴식 강제 금지', '갈 수 있는 길이 전부 휴식이 되지 않게'],
 ['규칙 2', '보스 직전 층은 항상 휴식', ''],
 ['규칙 3', '갈림길 두 갈래는 항상 서로 다른 종류', '고정 층은 예외'],
 ['', '', ''],
 ['일반 전투 보상', f"{rw['battle']['choices']}개 중 택1", f"족보 {rw['battle']['pool']['category']}% / 주사위 {rw['battle']['pool']['dice']}%"],
 ['일반 보상 등급', ' / '.join(f"{k} {v}%" for k,v in rw['battle']['tierWeights'].items()), ''],
 ['정예 보상', f"{rw['elite']['choices']}개 중 택1 + 유물 확정", f"족보 {rw['elite']['pool']['category']}% / 주사위 {rw['elite']['pool']['dice']}%"],
 ['정예 보상 등급', ' / '.join(f"{k} {v}%" for k,v in rw['elite']['tierWeights'].items()), ''],
 ['정예 유물', '항상 1개 드랍', '10% 확률로 정예 등급 유물'],
 ['보스 보상', '유물 + 전설 족보', ''],
 ['휴식 회복', f"최대 HP의 {int(act1['rest']['healRatio']*100)}%", '계몽 11단계부터 절반'],
]
sheet('지도·보상', '지도 생성 규칙 · 보상', '지도는 매 런 새로 만들어진다. 좌표를 직접 찍지 않고 층=행, 열=최대 4칸인 표로 배치한 뒤 브라우저가 정한 위치를 읽어 잇는다.',
      ['항목','값','비고'], rows, [22,34,44])

# ---------- 11. 경제 ----------
sh = act1['shop']; ds = act1['diceShift']
rows = [
 ['일반 전투 코인', f"{act1['coins']['battle'][0]}~{act1['coins']['battle'][1]}", ''],
 ['정예 전투 코인', f"{act1['coins']['elite'][0]}~{act1['coins']['elite'][1]}", ''],
 ['상점 주사위 진열', sh['stockDice'], ' / '.join(f"{k} {v}%" for k,v in sh['dieTierWeights'].items())],
 ['상점 유물 진열', sh['stockRelics'], ''],
 ['주사위 가격', ' / '.join(f"{k} {v}🪙" for k,v in sh['prices']['diceByTier'].items()), ''],
 ['유물 가격', f"{sh['prices']['relic']}🪙", ''],
 ['족보 교체 비용', f"변형당 {ds['perVariant']}🪙", f"처음 {ds['freeVariants']}개는 무료"],
 ['족보 교체 상한', f"보유 족보의 {int(ds['maxRatio']*100)}%", ''],
 ['계몽 13단계', '적이 주는 코인 -25%', ''],
 ['계몽 16단계', '상점 가격 증가', ''],
]
sheet('경제', '코인 · 상점 · 교체 비용', '코인은 전투에서만 나온다. 상점은 지도에 확률로 등장한다.',
      ['항목','값','비고'], rows, [22,34,34])

# ---------- 12. 계몽 ----------
ENL = ['엘리트가 더 자주 나타난다','모든 일반 적 공격력 +15%','모든 엘리트 공격력 +15%','보스 공격력 +15%',
 '보스 처치 회복 50% → 15%','HP 30%를 잃은 채 시작','모든 일반 적 체력 +20%','모든 엘리트 체력 +20%',
 '모든 보스 체력 +20%','주사위 하나가 저주 주사위로','휴식 회복량 -50%','일반의 언커먼·엘리트의 레어 확률 절반',
 '적이 주는 코인 -25%','최대 HP -10%','이벤트의 대가가 가혹해진다','상점 가격 증가','일반 적에게 계몽 패턴',
 '엘리트에게 계몽 패턴','보스에게 계몽 패턴','최종 보스가 두 마리']
rows = [[i+1, t, '적 강화' if '+15%' in t or '+20%' in t else ('플레이어 약화' if 'HP' in t and '적' not in t else '규칙 변경')] for i, t in enumerate(ENL)]
sheet('계몽', '계몽 20단계', '보스를 3막까지 깨면 1단계씩 오른다. 한 번 오른 단계는 내려가지 않는다.',
      ['단계','효과','분류'], rows, [8,44,16])

wb.save('/home/claude/redhood/docs/REDHOOD_game_db.xlsx')
print('저장 완료 · 시트', len(wb.sheetnames), wb.sheetnames)
